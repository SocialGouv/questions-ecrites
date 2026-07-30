#!/usr/bin/env python3
"""Extract DGCS-decided allotments from Salomé's Excel by parsing the
"Commentaires" column for Lot markers (e.g. "Lot AN 15650", "Lot SENAT
14822"). Groups QE by lot id and writes a CSV.

Two lookup passes are done to match each row's `N° QE` + Source to
our internal id `<AN|SENAT>-<leg>-QE-<num>`:
  1. Source column ("SENAT" / "AN") + numeric part of `N° QE`
  2. If ambiguous, prefer the legislature matching the file name

Usage:
    poetry run python scripts/extract_dgcs_lots.py \\
        --xlsx "data/1 - TABLEAU QE XVII LEG_Maj 09_10_2025 - vJD.xlsx" \\
        --leg 17 \\
        --output data/dgcs_lots_leg17.csv
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

# Matches "Lot AN 15650", "Lot Sénat 5029", "Lot QE AN 4034",
# "Lot SENAT 02130" (with leading zeros), etc.
LOT_RE = re.compile(
    r"\bLot\s*(?:QE\s+)?(AN|S[EÉ]NAT|SEN)\.?\s*(\d{2,7})\b",
    re.IGNORECASE,
)


def _normalise_source(cell: object) -> str | None:
    if cell is None:
        return None
    s = str(cell).strip().upper()
    if s.startswith("SENAT"):
        return "SENAT"
    if s.startswith("AN"):
        return "AN"
    return None


def _extract_num(cell: object) -> str | None:
    """From a cell like 'SENAT 127' or 'AN 12345', return the numeric
    part as a string. Returns None if no digits found."""
    if cell is None:
        return None
    s = str(cell).strip()
    # Some cells have hyperlink formatting or leading zeros — grab the
    # last run of digits.
    m = re.search(r"(\d{1,7})\b\s*$", s)
    if m:
        return m.group(1)
    m = re.search(r"(\d{1,7})", s)
    return m.group(1) if m else None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, required=True)
    ap.add_argument("--leg", type=int, required=True,
                    help="Législature (14/15/16/17) — used to build "
                         "internal ids like AN-<leg>-QE-<num>.")
    ap.add_argument("--output", type=Path, required=True)
    ap.add_argument("--header-row", type=int, default=4)
    ap.add_argument("--sheet", type=str, default=None,
                    help="Sheet name (defaults to first sheet).")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active
    header = list(ws.iter_rows(min_row=args.header_row,
                               max_row=args.header_row,
                               values_only=True))[0]

    def find_col(needles: list[str]) -> int:
        """1-indexed. Match by substring, case-insensitive."""
        for i, v in enumerate(header, start=1):
            if v is None:
                continue
            s = str(v).lower()
            for needle in needles:
                if needle in s:
                    return i
        raise KeyError(f"None of {needles} found in header")

    col_num       = find_col(["n° qe", "n° qe"])
    col_source    = find_col(["source"])
    col_pub       = find_col(["date de publication au jo"])
    col_objet     = find_col(["objet"])
    col_rep_date  = find_col(["date de réponse publiée",
                              "date de reponse publiee"])
    col_comments  = find_col(["commentaire"])

    lots: dict[str, list[dict]] = defaultdict(list)
    non_lot_rows = 0
    total_data_rows = 0
    missed_num = 0

    for row in ws.iter_rows(min_row=args.header_row + 1, values_only=True):
        # Skip fully-empty rows
        if not any(row):
            continue
        total_data_rows += 1

        source = _normalise_source(row[col_source - 1])
        num = _extract_num(row[col_num - 1])
        if not source or not num:
            missed_num += 1
            continue
        internal_id = f"{source}-{args.leg}-QE-{num}"

        comments = row[col_comments - 1]
        m = LOT_RE.search(str(comments) if comments else "")
        if not m:
            non_lot_rows += 1
            continue
        lot_source = m.group(1).upper()
        lot_source = "SENAT" if lot_source.startswith("SEN") else "AN"
        lot_num = m.group(2)
        lot_id = f"LOT-{lot_source}-{lot_num}"

        objet = row[col_objet - 1] or ""
        pub = row[col_pub - 1]
        rep = row[col_rep_date - 1]
        lots[lot_id].append({
            "lot_id": lot_id,
            "question_id": internal_id,
            "source_from_xlsx": source,
            "objet": str(objet).replace("\n", " ").strip(),
            "date_pub": pub.date().isoformat() if pub else "",
            "date_rep": rep.date().isoformat() if rep else "",
            "commentaires_raw": str(comments) if comments else "",
        })
    wb.close()

    groups_geq2 = {k: v for k, v in lots.items() if len(v) >= 2}
    sizes = Counter(len(v) for v in lots.values())

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "lot_id", "question_id", "source_from_xlsx", "objet",
            "date_pub", "date_rep", "commentaires_raw",
        ])
        w.writeheader()
        for lot in sorted(groups_geq2):
            for row in groups_geq2[lot]:
                w.writerow(row)

    print(f"Total data rows scanned : {total_data_rows}")
    print(f"Rows without Source/num : {missed_num}")
    print(f"Rows without Lot mention: {non_lot_rows}")
    print(f"Lots found              : {len(lots)}")
    print(f"Lots with >= 2 QE       : {len(groups_geq2)}")
    print(f"Distribution of lot sizes:")
    for size, cnt in sorted(sizes.items()):
        print(f"  {size:>3} QE : {cnt:>3} lots")
    print(f"\nWrote {args.output}")


if __name__ == "__main__":
    main()
